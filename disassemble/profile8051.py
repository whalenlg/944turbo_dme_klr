#!/usr/bin/env python3
"""
8051 Profiler — reads symbols from disassemble2.xlsx, sheet ASM, columns A-E.
"""
import re, sys, argparse, csv, os
from collections import defaultdict

def load_excel(path):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: pip install openpyxl --break-system-packages"); sys.exit(1)
    if not os.path.exists(path):
        print(f"ERROR: not found: {path}"); sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'ASM' not in wb.sheetnames:
        print(f"ERROR: no ASM sheet. Sheets: {wb.sheetnames}"); sys.exit(1)
    ws = wb['ASM']
    sorted_labels, addr_to_info = [], {}
    addr_re = re.compile(r'0x([0-9a-fA-F]+):?', re.IGNORECASE)
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        a_addr, b_label, c_opc, d_mnem, e_ops = row
        if a_addr is None: continue
        m = addr_re.match(str(a_addr).strip())
        if not m: continue
        addr = int(m.group(1), 16)
        label = None
        if b_label is not None:
            label = str(b_label).strip().rstrip(':').strip() or None
        opcodes = str(c_opc).strip() if c_opc is not None and not isinstance(c_opc, int) \
                  else (f"{c_opc:02x}" if isinstance(c_opc, int) else '')
        mnem = str(d_mnem).strip() if d_mnem is not None else ''
        ops  = str(e_ops).strip()  if e_ops  is not None else ''
        addr_to_info[addr] = {'label': label, 'mnem': mnem, 'ops': ops, 'opcodes': opcodes}
        if label:
            sorted_labels.append((addr, label))
    sorted_labels.sort()
    print(f"  Loaded {len(addr_to_info):,} addresses, {len(sorted_labels):,} labels from {path}")
    return sorted_labels, addr_to_info

def func_name_for(addr, sorted_labels):
    if not sorted_labels: return f"0x{addr:04X}"
    lo, hi, result = 0, len(sorted_labels)-1, None
    while lo <= hi:
        mid = (lo+hi)//2
        if sorted_labels[mid][0] <= addr: result=mid; lo=mid+1
        else: hi=mid-1
    return sorted_labels[result][1] if result is not None else f"0x{addr:04X}"

EXEC_RE = re.compile(r'\[EXEC\]\s+cyc=(\d+)\s+pc=([0-9a-fA-F]+)h\s+ir=([0-9a-fA-F]+)h', re.I)

def parse_trace(path):
    records = []
    with open(path, encoding='latin-1') as f:
        for line in f:
            m = EXEC_RE.search(line)
            if m:
                records.append((int(m.group(1)), int(m.group(2),16), int(m.group(3),16)))
    return records

def build_profile(records, sorted_labels):
    func_cycles, func_calls, pc_cycles, prev_func = defaultdict(int), defaultdict(int), defaultdict(int), None
    for i, (cyc, pc, ir) in enumerate(records):
        fname = func_name_for(pc, sorted_labels)
        if fname != prev_func:
            func_calls[fname] += 1; prev_func = fname
        elapsed = (records[i+1][0] - cyc) if i+1 < len(records) else 12
        func_cycles[fname] += elapsed
        pc_cycles[pc]      += elapsed
    return func_cycles, func_calls, pc_cycles, sum(func_cycles.values())

def fmt_time(cycles, mhz):
    ns = cycles / mhz * 1000
    if ns < 1e3:   return f"{ns:8.1f} ns"
    if ns < 1e6:   return f"{ns/1e3:8.3f} µs"
    return             f"{ns/1e6:8.3f} ms"

def print_report(func_cycles, func_calls, pc_cycles, total, sorted_labels, addr_to_info, args):
    mhz = args.clk_mhz
    entries = sorted(func_cycles.items(),
                     key=lambda x: (func_cycles[x[0]] if args.sort=='cycles'
                                    else func_calls[x[0]] if args.sort=='calls'
                                    else func_cycles[x[0]]/max(func_calls[x[0]],1)),
                     reverse=True)
    entries = [(n,c) for n,c in entries if c/total*100 >= args.min_pct][:args.top]

    W = 92
    print(); print("="*W)
    print("  8051 CYCLE PROFILE")
    print(f"  Total cycles : {total:,}    Wall time : {fmt_time(total,mhz).strip()}    @ {mhz} MHz")
    print("="*W)
    if args.calls:
        print(f"  {'Function':<28} {'Cycles':>13} {'%Total':>7} {'Wall Time':>13} {'Calls':>8} {'Cyc/Call':>10}")
    else:
        print(f"  {'Function':<28} {'Cycles':>13} {'%Total':>7} {'Wall Time':>13}")
    print("-"*W)

    shown = 0
    for fname, cycles in entries:
        pct = cycles/total*100; wt = fmt_time(cycles,mhz)
        calls = func_calls[fname]; avg = cycles/max(calls,1); shown += cycles
        if args.calls:
            print(f"  {fname:<28} {cycles:>13,} {pct:>6.2f}%  {wt} {calls:>8,} {avg:>10.1f}")
        else:
            print(f"  {fname:<28} {cycles:>13,} {pct:>6.2f}%  {wt}")

    print("-"*W)
    sp = shown/total*100
    print(f"  {'[shown above]':<28} {shown:>13,} {sp:>6.2f}%")
    rest = total-shown
    if rest > 0:
        print(f"  {'[other / filtered]':<28} {rest:>13,} {100-sp:>6.2f}%")
    print("="*W); print()

    if args.focus is not None:
        fa = int(args.focus,16); ff = func_name_for(fa, sorted_labels)
        ft = func_cycles.get(ff,1)
        print(f"  Per-PC breakdown: {ff}  ({ft:,} cycles total)")
        print("-"*W)
        print(f"  {'Addr':>6}  {'Cycles':>10}  {'%Func':>6}  {'Opcodes':<10}  {'Mnem':<8}  Operands")
        print("-"*W)
        items = sorted([(pc,c) for pc,c in pc_cycles.items()
                         if func_name_for(pc,sorted_labels)==ff], key=lambda x: x[0])
        for pc, cyc in items:
            pct = cyc/ft*100; info = addr_to_info.get(pc,{})
            lbl = info.get('label',''); lbl_s = f"  ← {lbl}" if lbl else ""
            print(f"  {pc:04X}h  {cyc:>10,}  {pct:>5.1f}%  "
                  f"{info.get('opcodes',''):<10}  {info.get('mnem',''):<8}  "
                  f"{info.get('ops','')}{lbl_s}")
        print("-"*W); print()

def write_csv(path, func_cycles, func_calls, total, mhz):
    with open(path,'w',newline='') as f:
        w = csv.writer(f)
        w.writerow(['function','cycles','pct','wall_ns','calls','cyc_per_call'])
        for fn, cyc in sorted(func_cycles.items(), key=lambda x: x[1], reverse=True):
            calls = func_calls[fn]
            w.writerow([fn, cyc, f"{cyc/total*100:.4f}", f"{cyc/mhz*1000:.1f}",
                        calls, f"{cyc/max(calls,1):.1f}"])
    print(f"  CSV written: {path}")

def main():
    ap = argparse.ArgumentParser(description="8051 cycle-accurate profiler")
    ap.add_argument('trace',  help="Simulation log with [EXEC] lines")
    ap.add_argument('excel', nargs='?', default=None,
                    help="disassemble2.xlsx (ASM sheet, cols A-E)")
    ap.add_argument('--top',     type=int,   default=30,   metavar='N')
    ap.add_argument('--min-pct', type=float, default=0.1,  metavar='F')
    ap.add_argument('--csv',     type=str,   default=None, metavar='FILE')
    ap.add_argument('--calls',   action='store_true')
    ap.add_argument('--sort',    choices=['cycles','calls','avg'], default='cycles')
    ap.add_argument('--focus',   type=str,   default=None, metavar='ADDR')
    ap.add_argument('--clk-mhz', type=float, default=6.0,  metavar='F')
    args = ap.parse_args()

    sorted_labels, addr_to_info = [], {}
    if args.excel:
        sorted_labels, addr_to_info = load_excel(args.excel)
    else:
        print("  No Excel file — addresses shown as hex")

    print(f"  Parsing trace: {args.trace} ...", end=' ', flush=True)
    records = parse_trace(args.trace)
    print(f"{len(records):,} EXEC events")
    if not records:
        print("ERROR: no [EXEC] lines found. Redirect vvp stdout: vvp sim > trace.log")
        sys.exit(1)

    func_cycles, func_calls, pc_cycles, total = build_profile(records, sorted_labels)
    print_report(func_cycles, func_calls, pc_cycles, total, sorted_labels, addr_to_info, args)
    if args.csv:
        write_csv(args.csv, func_cycles, func_calls, total, args.clk_mhz)

if __name__ == '__main__':
    main()
